"""버스 정류장 마스터 동기화.

입력: backend/app/batch/data/서울시버스노선별정류소정보(20260804).xlsx
  (ROUTE_ID/노선명/순번/NODE_ID/ARS_ID/정류소명/X좌표/Y좌표, 41,676행=노선×정류장 단위).
  NODE_ID(표준정류장ID)가 이 프로젝트 전체에서 쓰는 정류장 식별자다 — odsay_result.json의
  localStationID와 완전히 일치함을 확인했다(backend.md §7.3). X좌표=경도(lng),
  Y좌표=위도(lat) 순서인 점 주의(표준과 반대로 헷갈리기 쉬움, 실제 값 범위로 확인:
  X는 126~127대=경도, Y는 37대=위도).

노선×정류장 단위 원본에서 NODE_ID로 중복 제거하면 정류장 자체는 12,898개다(검증:
NODE_ID별로 정류소명·좌표가 전부 유일해 이름/좌표 충돌 없음, 노선이 여러 개 지나가는
정류장은 당연히 여러 행에 나타나지만 값은 같다).

출력: bus_stop 테이블에 UPSERT(stop_std_id = NODE_ID 기준).

사용법 (배치 러너에서):
    from app.batch.bus_stop_sync import sync_bus_stop_master
    n = sync_bus_stop_master(cur)
"""
from pathlib import Path

import openpyxl

_DATA = Path(__file__).parent / "data" / "서울시버스노선별정류소정보(20260804).xlsx"

_UPSERT_SQL = """
    INSERT INTO bus_stop (stop_std_id, stop_name, lat, lng)
    VALUES (%(stop_std_id)s, %(stop_name)s, %(lat)s, %(lng)s)
    ON CONFLICT (stop_std_id)
    DO UPDATE SET stop_name = EXCLUDED.stop_name, lat = EXCLUDED.lat, lng = EXCLUDED.lng
"""


def _load_rows() -> list[dict]:
    wb = openpyxl.load_workbook(_DATA, read_only=True, data_only=True)
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(header)}

    seen: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        node_id = str(row[idx["NODE_ID"]])
        if node_id in seen:
            continue
        seen[node_id] = {
            "stop_std_id": node_id,
            "stop_name": row[idx["정류소명"]],
            "lng": float(row[idx["X좌표"]]),
            "lat": float(row[idx["Y좌표"]]),
        }
    wb.close()
    return list(seen.values())


def sync_bus_stop_master(cur) -> int:
    """bus_stop 테이블을 정류소정보 파일 기준으로 UPSERT한다. 반영된 행 수를 반환한다."""
    rows = _load_rows()
    cur.executemany(_UPSERT_SQL, rows)
    return len(rows)
